import openpyxl
import os
import sys

files = [f for f in os.listdir('.') if f.endswith('.xlsx')]
if not files:
    print("No xlsx found")
    sys.exit(1)

wb = openpyxl.load_workbook(files[0], read_only=True, data_only=True)

output_lines = []
output_lines.append("=== SHEET NAMES ===")
for name in wb.sheetnames:
    output_lines.append(f"  [{name}]")

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    output_lines.append("")
    output_lines.append("=" * 80)
    output_lines.append(f"SHEET: {sheet_name}")
    output_lines.append("=" * 80)
    row_count = 0
    for row in ws.iter_rows(max_row=200, values_only=True):
        non_empty = []
        for c in row:
            v = str(c).strip() if c is not None else ''
            if v and v != 'None':
                non_empty.append(v)
        if non_empty:
            output_lines.append(" | ".join(non_empty[:12]))
            row_count += 1
        if row_count >= 150:
            output_lines.append("... (row limit reached)")
            break

wb.close()

with open('excel_output.txt', 'w', encoding='utf-8') as f:
    f.write('\n'.join(output_lines))

print(f"Done! Written {len(output_lines)} lines to excel_output.txt")
